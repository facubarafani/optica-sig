"""Bulk import: readers, column mapping, preview and atomic commit."""
import io

import pytest

from app.services.importer.readers import parse_decimal


# --- helpers ---------------------------------------------------------------

def xlsx(rows: list[list]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload(client, headers, spec, content, filename="datos.xlsx"):
    return client.post(
        f"/api/imports/{spec}/upload",
        files={"file": (filename, content, "application/octet-stream")},
        headers=headers,
    )


def run_import(client, headers, spec, rows, *, filename="datos.xlsx",
               content=None, mapping=None, decimal_format="es",
               create_missing=True, expect_preview_ok=True):
    """Upload → preview → commit, returning (preview, result)."""
    body = content if content is not None else xlsx(rows)
    up = upload(client, headers, spec, body, filename)
    assert up.status_code == 200, up.text
    up = up.json()
    opts = {
        "mapping": mapping or up["suggested_mapping"],
        "decimal_format": decimal_format,
        "create_missing": create_missing,
    }
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json=opts, headers=headers)
    assert pv.status_code == 200, pv.text
    preview = pv.json()
    if not expect_preview_ok:
        return preview, None
    assert preview["ok"], preview["errors"]
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=headers)
    assert res.status_code == 200, res.text
    return preview, res.json()


PRODUCT_HEADERS = ["Código", "Descripción", "Tipo de producto", "Marca",
                   "Modelo", "Costo"]


# --- number parsing --------------------------------------------------------

@pytest.mark.parametrize("raw, fmt, expected", [
    ("12.500,50", "es", "12500.50"),
    ("12500,50", "es", "12500.50"),
    ("1.234.567,89", "es", "1234567.89"),
    ("$ 20.000", "es", "20000"),
    ("12,500.50", "en", "12500.50"),
    ("12500.50", "en", "12500.50"),
])
def test_parse_decimal_text(raw, fmt, expected):
    assert str(parse_decimal(raw, decimal_format=fmt)) == expected


def test_numeric_cells_are_never_locale_parsed():
    # A real number out of Excel is already unambiguous. Applying es-AR rules to
    # "12500.5" would strip the dot and yield 125005 — a 10x error.
    assert str(parse_decimal(12500.5, decimal_format="es")) == "12500.5"
    assert str(parse_decimal(10, decimal_format="es")) == "10"


def test_parse_decimal_rejects_garbage():
    with pytest.raises(ValueError):
        parse_decimal("12.5O0", decimal_format="es")


# --- specs & template ------------------------------------------------------

def test_specs_are_listed(client, auth_headers):
    resp = client.get("/api/imports/specs", headers=auth_headers)
    assert resp.status_code == 200
    keys = {s["key"] for s in resp.json()}
    assert keys == {"products", "costs", "stock", "price_list_items"}


def test_template_is_a_real_xlsx(client, auth_headers):
    resp = client.get("/api/imports/products/template", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"          # xlsx is a zip
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Referencia" in wb.sheetnames
    assert wb["Productos"].cell(row=1, column=1).value == "Código"


# --- products --------------------------------------------------------------

def test_import_creates_products_and_missing_catalog_entries(
    client, auth_headers, product_type_id
):
    rows = [
        PRODUCT_HEADERS,
        ["ARM-100", "Armazón negro", "Armazones", "Vulk", "Clipper", "20.000,00"],
        ["ARM-101", "Armazón rojo", "Armazones", "Vulk", "Aviador", "21.500,50"],
    ]
    preview, result = run_import(client, auth_headers, "products", rows)
    assert (preview["to_create"], preview["to_update"]) == (2, 0)
    assert {m["name"] for m in preview["missing_refs"]} == {"Vulk", "Clipper", "Aviador"}
    assert result["created"] == 2

    listed = client.get("/api/products", headers=auth_headers).json()
    by_code = {p["code"]: p for p in listed}
    assert by_code["ARM-100"]["current_cost"] == "20000.00"
    assert by_code["ARM-101"]["current_cost"] == "21500.50"
    assert by_code["ARM-100"]["brand_id"] == by_code["ARM-101"]["brand_id"]

    brands = client.get("/api/brands", headers=auth_headers).json()
    assert [b["name"] for b in brands] == ["Vulk"]


def test_reimport_updates_instead_of_duplicating(
    client, auth_headers, product_type_id
):
    rows = [PRODUCT_HEADERS, ["ARM-100", "Original", "Armazones", "", "", "100"]]
    run_import(client, auth_headers, "products", rows)

    rows2 = [PRODUCT_HEADERS, ["ARM-100", "Corregido", "Armazones", "", "", "150"]]
    preview, result = run_import(client, auth_headers, "products", rows2)
    assert (preview["to_create"], preview["to_update"]) == (0, 1)
    assert result["updated"] == 1

    listed = client.get("/api/products", headers=auth_headers).json()
    assert len([p for p in listed if p["code"] == "ARM-100"]) == 1
    p = next(p for p in listed if p["code"] == "ARM-100")
    assert p["description"] == "Corregido"
    assert p["current_cost"] == "150.00"
    # An updated cost must be historised, not silently overwritten.
    hist = client.get(f"/api/products/{p['id']}/cost-history",
                      headers=auth_headers).json()
    assert [h["new_cost"] for h in hist] == ["150.00"]


def test_preview_reports_every_bad_row_and_writes_nothing(
    client, auth_headers, product_type_id
):
    rows = [
        PRODUCT_HEADERS,
        ["ARM-200", "ok", "Armazones", "", "", "1000"],
        ["ARM-201", "costo roto", "Armazones", "", "", "12.5O0"],
        ["", "sin código", "Armazones", "", "", "10"],
        ["ARM-200", "repetido", "Armazones", "", "", "10"],
    ]
    preview, _ = run_import(client, auth_headers, "products", rows,
                            expect_preview_ok=False)
    assert not preview["ok"]
    messages = " ".join(e["message"] for e in preview["errors"])
    assert "no es un número válido" in messages
    assert "Falta" in messages
    assert "Repetido en el archivo" in messages
    assert client.get("/api/products", headers=auth_headers).json() == []


def test_commit_is_all_or_nothing(client, auth_headers, product_type_id):
    rows = [
        PRODUCT_HEADERS,
        ["ARM-300", "ok", "Armazones", "", "", "1000"],
        ["ARM-301", "roto", "Armazones", "", "", "no-es-numero"],
    ]
    up = upload(client, auth_headers, "products", xlsx(rows)).json()
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": True}
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers)
    assert res.status_code == 400
    # The good row must not have landed either.
    assert client.get("/api/products", headers=auth_headers).json() == []


def test_unknown_supplier_is_an_error_not_an_autocreate(
    client, auth_headers, product_type_id
):
    rows = [
        ["Código", "Tipo de producto", "Proveedor"],
        ["ARM-400", "Armazones", "Proveedor Fantasma"],
    ]
    preview, _ = run_import(client, auth_headers, "products", rows,
                            expect_preview_ok=False)
    assert any("no existe" in e["message"] for e in preview["errors"])


def test_create_missing_false_refuses_to_invent_catalog_entries(
    client, auth_headers, product_type_id
):
    rows = [PRODUCT_HEADERS, ["ARM-500", "x", "Armazones", "MarcaNueva", "", "10"]]
    up = upload(client, auth_headers, "products", xlsx(rows)).json()
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": False}
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers)
    assert res.status_code == 400
    assert "MarcaNueva" in res.json()["detail"]


def test_committed_batch_cannot_be_replayed(client, auth_headers, product_type_id):
    rows = [PRODUCT_HEADERS, ["ARM-600", "x", "Armazones", "", "", "10"]]
    up = upload(client, auth_headers, "products", xlsx(rows)).json()
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": True}
    client.post(f"/api/imports/batches/{up['batch_id']}/preview", json=opts,
                headers=auth_headers)
    first = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                        json=opts, headers=auth_headers)
    assert first.status_code == 200
    again = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                        json=opts, headers=auth_headers)
    assert again.status_code == 409


# --- other file shapes -----------------------------------------------------

def test_csv_with_semicolons_and_argentine_numbers(
    client, auth_headers, product_type_id
):
    csv_bytes = (
        "Código;Descripción;Tipo de producto;Costo\n"
        "ARM-700;Armazón ñandú;Armazones;12.500,50\n"
    ).encode("utf-8")
    _, result = run_import(client, auth_headers, "products", None,
                           content=csv_bytes, filename="lista.csv")
    assert result["created"] == 1
    p = client.get("/api/products", headers=auth_headers).json()[0]
    assert p["current_cost"] == "12500.50"
    assert p["description"] == "Armazón ñandú"      # accents survived


def test_numeric_excel_cells_keep_their_value(client, auth_headers, product_type_id):
    # Written as real numbers, not text — the classic 12500.5 -> 125005 trap.
    rows = [["Código", "Tipo de producto", "Costo"], ["ARM-800", "Armazones", 12500.5]]
    _, result = run_import(client, auth_headers, "products", rows)
    assert result["created"] == 1
    assert client.get("/api/products",
                      headers=auth_headers).json()[0]["current_cost"] == "12500.50"


def test_supplier_headers_can_be_remapped(client, auth_headers, product_type_id):
    rows = [
        ["ARTICULO", "RUBRO", "COSTO S/IVA", "OBS"],
        ["ARM-900", "Armazones", "5.000", "ignorar esto"],
    ]
    up = upload(client, auth_headers, "products", xlsx(rows)).json()
    assert up["suggested_mapping"] == {}         # nothing matches automatically
    mapping = {"code": "ARTICULO", "product_type": "RUBRO",
               "current_cost": "COSTO S/IVA"}
    opts = {"mapping": mapping, "decimal_format": "es", "create_missing": True}
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json=opts, headers=auth_headers).json()
    assert pv["ok"], pv["errors"]
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers)
    assert res.status_code == 200
    assert client.get("/api/products",
                      headers=auth_headers).json()[0]["current_cost"] == "5000.00"


def test_missing_required_column_is_rejected_up_front(
    client, auth_headers, product_type_id
):
    rows = [["Descripción"], ["sin código ni tipo"]]
    up = upload(client, auth_headers, "products", xlsx(rows)).json()
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json={"mapping": {}, "decimal_format": "es"},
                     headers=auth_headers)
    assert pv.status_code == 400
    assert "obligatorias" in pv.json()["detail"]


def test_unsupported_format_is_refused(client, auth_headers):
    resp = upload(client, auth_headers, "products", b"%PDF-1.4 fake", "lista.pdf")
    assert resp.status_code == 400
    assert "pdf" in resp.json()["detail"].lower()


# --- costs / stock / price lists ------------------------------------------

def test_cost_import_historises_each_change(client, auth_headers, product_id):
    rows = [["Código", "Costo nuevo", "Nota"], ["P-1", "250,50", "lista julio"]]
    _, result = run_import(client, auth_headers, "costs", rows)
    assert result["updated"] == 1

    hist = client.get(f"/api/products/{product_id}/cost-history",
                      headers=auth_headers).json()
    assert hist[0]["new_cost"] == "250.50"
    assert hist[0]["old_cost"] == "100.00"
    assert hist[0]["note"] == "lista julio"


def test_stock_import_sets_the_level_and_is_idempotent(
    client, auth_headers, product_id, branch_id
):
    rows = [["Código", "Sucursal", "Cantidad"], ["P-1", "Central", "10"]]
    _, result = run_import(client, auth_headers, "stock", rows)
    assert result["updated"] == 1
    levels = client.get(f"/api/stock/levels?product_id={product_id}",
                        headers=auth_headers).json()
    assert levels[0]["quantity"] == "10.00"

    # Re-importing the same file must not double the stock.
    _, result2 = run_import(client, auth_headers, "stock", rows)
    assert result2["updated"] == 0
    levels = client.get(f"/api/stock/levels?product_id={product_id}",
                        headers=auth_headers).json()
    assert levels[0]["quantity"] == "10.00"

    # And it writes a real movement, never a direct level mutation.
    movements = client.get("/api/stock/movements", headers=auth_headers).json()
    assert [m["movement_type"] for m in movements] == ["adjustment"]


def test_price_list_items_import(client, auth_headers):
    """A code the list does not have yet is appended to its ladder."""
    lst = client.post("/api/price-lists", json={"name": "Lista General"},
                      headers=auth_headers).json()
    rows = [["Lista de precios", "Categoría", "Precio"],
            ["Lista General", "AA", "50.000,00"]]
    _, result = run_import(client, auth_headers, "price_list_items", rows)
    assert result["created"] == 1

    cats = client.get(f"/api/price-lists/{lst['id']}/categories",
                      headers=auth_headers).json()
    assert cats[0]["code"] == "AA"
    assert cats[0]["price"] == "50000.00"


def test_price_list_items_import_updates_an_existing_code(client, auth_headers):
    lst = client.post("/api/price-lists", json={"name": "Lista General"},
                      headers=auth_headers).json()
    client.post(f"/api/price-lists/{lst['id']}/generate-categories",
                json={"count": 2}, headers=auth_headers)

    rows = [["Lista de precios", "Categoría", "Precio"],
            ["Lista General", "AB", "77.000,00"]]
    preview, result = run_import(client, auth_headers, "price_list_items", rows)
    assert preview["to_update"] == 1 and preview["to_create"] == 0
    assert result["updated"] == 1

    cats = client.get(f"/api/price-lists/{lst['id']}/categories",
                      headers=auth_headers).json()
    assert [c["price"] for c in cats] == ["0.00", "77000.00"]


def test_batches_are_recorded_in_the_history(client, auth_headers, product_type_id):
    rows = [PRODUCT_HEADERS, ["ARM-950", "x", "Armazones", "", "", "10"]]
    run_import(client, auth_headers, "products", rows)
    batches = client.get("/api/imports/batches", headers=auth_headers).json()
    assert len(batches) == 1
    assert batches[0]["status"] == "committed"
    assert batches[0]["spec_key"] == "products"
    assert batches[0]["row_count"] == 1
