"""Export mirrors import: same columns, and the output re-imports cleanly."""
import io

import pytest


def get_export(client, headers, spec, fmt="xlsx", **params):
    qs = "&".join(f"{k}={v}" for k, v in params.items())
    url = f"/api/imports/{spec}/export?format={fmt}" + (f"&{qs}" if qs else "")
    return client.get(url, headers=headers)


def read_xlsx(content):
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(content)).active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


@pytest.fixture
def catalogued_product(client, auth_headers, product_type_id):
    brand = client.post("/api/brands", json={"name": "Vulk"}, headers=auth_headers).json()
    model = client.post("/api/product-models", json={"name": "Clipper"},
                        headers=auth_headers).json()
    resp = client.post("/api/products", json={
        "code": "ARM-001", "description": "Armazón clásico",
        "product_type_id": product_type_id, "brand_id": brand["id"],
        "model_id": model["id"], "current_cost": "20000.50", "min_stock": "2",
    }, headers=auth_headers)
    assert resp.status_code == 201, resp.text
    return resp.json()


# --- shape -----------------------------------------------------------------

def test_headers_match_the_import_template(client, auth_headers, catalogued_product):
    tpl = client.get("/api/imports/products/template", headers=auth_headers)
    tpl_headers, _ = read_xlsx(tpl.content)

    exp = get_export(client, auth_headers, "products")
    assert exp.status_code == 200, exp.text
    exp_headers, rows = read_xlsx(exp.content)

    assert list(exp_headers) == list(tpl_headers)
    assert len(rows) == 1


def test_xlsx_writes_real_numbers_not_text(client, auth_headers, catalogued_product):
    exp = get_export(client, auth_headers, "products")
    headers, rows = read_xlsx(exp.content)
    cost = rows[0][list(headers).index("Costo")]
    assert isinstance(cost, (int, float)), f"got {type(cost)}"
    assert float(cost) == 20000.50


def test_csv_uses_semicolons_and_comma_decimals(client, auth_headers, catalogued_product):
    exp = get_export(client, auth_headers, "products", fmt="csv")
    assert exp.status_code == 200
    assert exp.content.startswith(b"\xef\xbb\xbf")        # BOM for Excel
    text = exp.content.decode("utf-8-sig")
    assert ";" in text.splitlines()[0]
    assert "20000,50" in text
    assert "Armazón clásico" in text


def test_unknown_format_is_rejected(client, auth_headers):
    resp = get_export(client, auth_headers, "products", fmt="pdf")
    assert resp.status_code == 400
    assert "pdf" in resp.json()["detail"]


def test_every_import_spec_can_be_exported(client, auth_headers):
    for spec in ("products", "costs", "stock", "price_list_items"):
        for fmt in ("xlsx", "csv"):
            resp = get_export(client, auth_headers, spec, fmt=fmt)
            assert resp.status_code == 200, f"{spec}/{fmt}: {resp.text}"


# --- filters ---------------------------------------------------------------

def test_export_respects_filters_and_inactive(
    client, auth_headers, product_type_id, catalogued_product
):
    other_type = client.post("/api/product-types", json={"name": "Líquidos"},
                             headers=auth_headers).json()["id"]
    client.post("/api/products",
                json={"code": "LIQ-1", "product_type_id": other_type},
                headers=auth_headers)

    _, rows = read_xlsx(get_export(client, auth_headers, "products").content)
    assert len(rows) == 2

    _, rows = read_xlsx(get_export(client, auth_headers, "products",
                                   product_type_id=product_type_id).content)
    assert [r[0] for r in rows] == ["ARM-001"]

    # Deactivated rows are out unless asked for.
    client.delete(f"/api/products/{catalogued_product['id']}", headers=auth_headers)
    _, rows = read_xlsx(get_export(client, auth_headers, "products").content)
    assert [r[0] for r in rows] == ["LIQ-1"]
    _, rows = read_xlsx(get_export(client, auth_headers, "products",
                                   include_inactive="true").content)
    assert sorted(r[0] for r in rows) == ["ARM-001", "LIQ-1"]


# --- the point of it: round trip -------------------------------------------

def test_xlsx_round_trip(client, auth_headers, catalogued_product):
    """Export, edit a cell, re-import: the product is updated, not duplicated."""
    exported = get_export(client, auth_headers, "products").content
    headers, rows = read_xlsx(exported)

    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(list(headers))
    row = list(rows[0])
    row[list(headers).index("Descripción")] = "Editado en Excel"
    row[list(headers).index("Costo")] = 31000.75
    ws.append(row)
    buf = io.BytesIO(); wb.save(buf)

    up = client.post("/api/imports/products/upload",
                     files={"file": ("editado.xlsx", buf.getvalue(),
                                     "application/octet-stream")},
                     headers=auth_headers).json()
    # Round trip means the mapping resolves itself — no manual column matching.
    assert up["suggested_mapping"]["code"] == "Código"
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": True}
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json=opts, headers=auth_headers).json()
    assert pv["ok"], pv["errors"]
    assert (pv["to_create"], pv["to_update"]) == (0, 1)
    assert pv["missing_refs"] == []       # brand/model already resolve by name

    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers)
    assert res.status_code == 200, res.text

    products = client.get("/api/products", headers=auth_headers).json()
    assert len(products) == 1
    assert products[0]["description"] == "Editado en Excel"
    assert products[0]["current_cost"] == "31000.75"


def test_csv_round_trip(client, auth_headers, catalogued_product):
    """The CSV we emit must re-import under the wizard's default 'es' format."""
    csv_bytes = get_export(client, auth_headers, "products", fmt="csv").content
    up = client.post("/api/imports/products/upload",
                     files={"file": ("export.csv", csv_bytes, "text/csv")},
                     headers=auth_headers).json()
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": True}
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json=opts, headers=auth_headers).json()
    assert pv["ok"], pv["errors"]
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers)
    assert res.status_code == 200, res.text
    p = client.get("/api/products", headers=auth_headers).json()[0]
    assert p["current_cost"] == "20000.50"     # decimals survived the trip
    assert p["description"] == "Armazón clásico"


def test_stock_export_round_trips(client, auth_headers, product_id, branch_id):
    client.post("/api/stock/movements", json={
        "product_id": product_id, "branch_id": branch_id,
        "movement_type": "inbound", "quantity": "7",
    }, headers=auth_headers)

    _, rows = read_xlsx(get_export(client, auth_headers, "stock").content)
    assert len(rows) == 1 and float(rows[0][2]) == 7.0

    csv_bytes = get_export(client, auth_headers, "stock", fmt="csv").content
    up = client.post("/api/imports/stock/upload",
                     files={"file": ("stock.csv", csv_bytes, "text/csv")},
                     headers=auth_headers).json()
    opts = {"mapping": up["suggested_mapping"], "decimal_format": "es",
            "create_missing": False}
    pv = client.post(f"/api/imports/batches/{up['batch_id']}/preview",
                     json=opts, headers=auth_headers).json()
    assert pv["ok"], pv["errors"]
    res = client.post(f"/api/imports/batches/{up['batch_id']}/commit",
                      json=opts, headers=auth_headers).json()
    # Re-importing the current state is a no-op, not a doubling.
    assert res["updated"] == 0
    levels = client.get(f"/api/stock/levels?product_id={product_id}",
                        headers=auth_headers).json()
    assert levels[0]["quantity"] == "7.00"


def test_export_requires_the_read_permission(client, auth_headers, db):
    """A user without stock:read cannot export stock."""
    from app.models.auth import User
    resp = client.post("/api/users", json={
        "email": "solo@lectura.com", "password": "secret1234",
        "full_name": "Sólo productos", "role_ids": [],
    }, headers=auth_headers)
    assert resp.status_code == 201, resp.text
    tok = client.post("/api/auth/login",
                      json={"email": "solo@lectura.com", "password": "secret1234"}).json()
    h = {"Authorization": f"Bearer {tok['access_token']}"}
    assert get_export(client, h, "stock").status_code == 403
