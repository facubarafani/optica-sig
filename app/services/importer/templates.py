"""Generate the .xlsx template for an import spec.

Two sheets: the one you fill in, and a "Referencia" sheet listing the values
that already exist in this company, so you are not guessing how a brand or a
branch is spelled.
"""
from __future__ import annotations

import io

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.branch import Branch
from app.models.pricing import PriceCategory, PriceList
from app.models.product import Brand, ProductModel, ProductType
from app.models.supplier import Supplier
from app.services.importer.specs import ImportSpec

_REFERENCE_SOURCES = [
    ("Tipos de producto", ProductType),
    ("Marcas", Brand),
    ("Modelos", ProductModel),
    ("Proveedores", Supplier),
    ("Sucursales", Branch),
    ("Listas de precios", PriceList),
]


def build_template(db: Session, spec: ImportSpec, *, company_id: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = spec.label[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    req_fill = PatternFill("solid", fgColor="1D4ED8")

    for col, f in enumerate(spec.fields, start=1):
        cell = ws.cell(row=1, column=col, value=f.label)
        cell.font = header_font
        cell.fill = req_fill if f.required else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if f.help or f.required:
            note = f.help or ""
            if f.required:
                note = ("OBLIGATORIO. " + note).strip()
            cell.comment = _comment(note)
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(f.label) + 4)

    # One example row so the expected shape is obvious. It is clearly marked so
    # nobody imports it by accident.
    for col, f in enumerate(spec.fields, start=1):
        ws.cell(row=2, column=col, value=f.example or "")
    ws.cell(row=2, column=1).font = Font(italic=True, color="9CA3AF")
    ws.cell(
        row=3, column=1,
        value="↑ La fila 2 es un ejemplo: borrala antes de importar.",
    ).font = Font(italic=True, color="DC2626")

    ws.freeze_panes = "A2"

    _reference_sheet(wb, db, company_id=company_id)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _comment(text: str):
    from openpyxl.comments import Comment

    c = Comment(text, "SGI Óptica")
    c.width, c.height = 260, 90
    return c


def _reference_sheet(wb, db: Session, *, company_id: int) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("Referencia")
    ws.cell(row=1, column=1, value="Valores válidos ya cargados en el sistema") \
        .font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value="Escribí los nombres exactamente así. Tipos, marcas y modelos "
                  "que no existan se pueden crear al importar.") \
        .font = Font(italic=True, color="6B7280")

    col = 1
    for label, model in _REFERENCE_SOURCES:
        names = list(
            db.execute(
                select(model.name)
                .where(model.company_id == company_id, model.is_active.is_(True))
                .order_by(model.name)
            ).scalars()
        )
        _reference_column(ws, col, label, names)
        col += 1

    # Categories belong to a list, so they are listed as "Lista · Código"
    # rather than as a flat set of names.
    steps = db.execute(
        select(PriceList.name, PriceCategory.code)
        .join(PriceCategory, PriceCategory.price_list_id == PriceList.id)
        .where(
            PriceCategory.company_id == company_id,
            PriceCategory.is_active.is_(True),
            PriceList.is_active.is_(True),
        )
        .order_by(PriceList.name, PriceCategory.position, PriceCategory.id)
    ).all()
    _reference_column(
        ws, col, "Categorías por lista", [f"{r[0]} · {r[1]}" for r in steps]
    )


def _reference_column(ws, col: int, label: str, values: list[str]) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws.cell(row=4, column=col, value=label).font = Font(bold=True)
    for i, value in enumerate(values, start=5):
        ws.cell(row=i, column=col, value=value)
    ws.column_dimensions[get_column_letter(col)].width = max(18, len(label) + 4)
