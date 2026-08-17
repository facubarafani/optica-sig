"""Export the same shapes we import — the mirror of engine.py.

Sharing ``specs.py`` with the importer is the whole point: an export uses the
exact column set the importer expects, so the round trip works. Export the
catalogue, edit it in Excel, upload it again and every row matches by code.

Conventions, chosen so the file re-imports cleanly with the wizard's defaults:
  * .xlsx  — numbers are written as real numeric cells, which the reader takes
             verbatim (no locale guessing at all).
  * .csv   — ';' separator and comma decimals, i.e. what Spanish Excel both
             produces and expects, matching the importer's default "es" format.
             A UTF-8 BOM is written so Excel keeps the accents.
"""
from __future__ import annotations

import csv
import io
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.branch import Branch
from app.models.pricing import PriceCategory, PriceList
from app.models.product import Brand, Color, Product, ProductModel, ProductType
from app.models.supplier import Supplier
from app.services import stock as stock_service
from app.services.importer.specs import ImportSpec


class ExportError(Exception):
    pass


def _name_map(db: Session, model, company_id: int) -> dict[int, str]:
    return {
        r[0]: r[1]
        for r in db.execute(
            select(model.id, model.name).where(model.company_id == company_id)
        ).all()
    }


# --- row builders (one per spec, keyed by field) ---------------------------

def _products(db: Session, company_id: int, opts: dict) -> list[dict]:
    types = _name_map(db, ProductType, company_id)
    brands = _name_map(db, Brand, company_id)
    models = _name_map(db, ProductModel, company_id)
    suppliers = _name_map(db, Supplier, company_id)
    colors = _name_map(db, Color, company_id)
    lists_ = _name_map(db, PriceList, company_id)

    stmt = select(Product).where(Product.company_id == company_id)
    if not opts.get("include_inactive"):
        stmt = stmt.where(Product.is_active.is_(True))
    for field, key in (
        ("product_type_id", "product_type_id"), ("brand_id", "brand_id"),
        ("model_id", "model_id"), ("supplier_id", "supplier_id"),
        ("color_id", "color_id"),
    ):
        if opts.get(key):
            stmt = stmt.where(getattr(Product, field) == opts[key])

    rows = []
    for p in db.execute(stmt.order_by(Product.code)).scalars():
        rows.append({
            "code": p.code,
            "description": p.description,
            "product_type": types.get(p.product_type_id),
            "brand": brands.get(p.brand_id),
            "model": models.get(p.model_id),
            "supplier": suppliers.get(p.supplier_id),
            "color": colors.get(p.color_id),
            "current_cost": p.current_cost,
            "min_stock": p.min_stock,
            # The enum *value*, which is what the import spec accepts.
            "pricing_mode": p.pricing_mode.value,
            "sale_price": p.sale_price,
            "price_list": lists_.get(p.price_list_id),
            "price_category": p.price_category_code,
        })
    return rows


def _costs(db: Session, company_id: int, opts: dict) -> list[dict]:
    stmt = select(Product).where(
        Product.company_id == company_id, Product.is_active.is_(True)
    )
    if opts.get("supplier_id"):
        stmt = stmt.where(Product.supplier_id == opts["supplier_id"])
    return [
        {"product": p.code, "new_cost": p.current_cost, "note": None}
        for p in db.execute(stmt.order_by(Product.code)).scalars()
    ]


def _stock(db: Session, company_id: int, opts: dict) -> list[dict]:
    products = {
        r[0]: r[1] for r in db.execute(
            select(Product.id, Product.code).where(Product.company_id == company_id)
        ).all()
    }
    branches = _name_map(db, Branch, company_id)
    # Delegated rather than re-queried: the Stock screen and this export must
    # agree on what the filters mean, so only one place defines it.
    levels = stock_service.list_levels(
        db,
        company_id=company_id,
        branch_id=opts.get("branch_id"),
        q=opts.get("q"),
        product_type_id=opts.get("product_type_id"),
        brand_id=opts.get("brand_id"),
        color_id=opts.get("color_id"),
        low_only=bool(opts.get("low_only")),
        include_inactive=bool(opts.get("include_inactive")),
    )
    rows = [
        {
            "product": products.get(lvl.product_id),
            "branch": branches.get(lvl.branch_id),
            "quantity": lvl.quantity,
        }
        for lvl in levels
    ]
    return [r for r in rows if r["product"] and r["branch"]]


def _price_list_items(db: Session, company_id: int, opts: dict) -> list[dict]:
    lists_ = _name_map(db, PriceList, company_id)
    stmt = select(PriceCategory).where(PriceCategory.company_id == company_id)
    if opts.get("price_list_id"):
        stmt = stmt.where(PriceCategory.price_list_id == opts["price_list_id"])
    if not opts.get("include_inactive"):
        stmt = stmt.where(PriceCategory.is_active.is_(True))
    rows = []
    # Ladder order, so the exported file reads cheapest-first like the console.
    order = (PriceCategory.price_list_id, PriceCategory.position, PriceCategory.id)
    for cat in db.execute(stmt.order_by(*order)).scalars():
        rows.append({
            "price_list": lists_.get(cat.price_list_id),
            "price_category": cat.code,
            "price": cat.price,
            "description": cat.description,
        })
    return [r for r in rows if r["price_list"]]


BUILDERS = {
    "products": _products,
    "costs": _costs,
    "stock": _stock,
    "price_list_items": _price_list_items,
}


def build_rows(db: Session, spec: ImportSpec, *, company_id: int, **opts) -> list[dict]:
    builder = BUILDERS.get(spec.key)
    if builder is None:
        raise ExportError(f"No hay exportación definida para '{spec.key}'.")
    return builder(db, company_id, opts)


# --- writers ---------------------------------------------------------------

def _cell(value):
    return "" if value is None else value


def to_xlsx(spec: ImportSpec, rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = spec.label[:31]
    for col, f in enumerate(spec.fields, start=1):
        c = ws.cell(row=1, column=col, value=f.label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8" if f.required else "2563EB")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(f.label) + 4)
    for i, row in enumerate(rows, start=2):
        for col, f in enumerate(spec.fields, start=1):
            # Decimals go in as numbers, not text: the reader then takes them
            # verbatim and no locale setting can corrupt them.
            ws.cell(row=i, column=col, value=_cell(row.get(f.key)))
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(spec: ImportSpec, rows: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow([f.label for f in spec.fields])
    for row in rows:
        out = []
        for f in spec.fields:
            v = row.get(f.key)
            if v is None:
                out.append("")
            elif isinstance(v, Decimal):
                out.append(str(v).replace(".", ","))   # es-AR decimal comma
            else:
                out.append(str(v))
        w.writerow(out)
    # BOM so Excel opens it as UTF-8 and keeps the accents.
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


FORMATS = {
    "xlsx": (
        to_xlsx,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
    "csv": (to_csv, "text/csv; charset=utf-8"),
}


def render(spec: ImportSpec, rows: list[dict], fmt: str) -> tuple[bytes, str]:
    if fmt not in FORMATS:
        raise ExportError(
            f"Formato '{fmt}' no soportado. Usá: {', '.join(FORMATS)}."
        )
    writer, media_type = FORMATS[fmt]
    return writer(spec, rows), media_type
