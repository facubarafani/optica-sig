"""Turn an uploaded file into (headers, rows).

This is the only module that knows about file formats. Everything downstream
works on plain JSON-serialisable cell values, so adding a format later (PDF
tables, ODS…) means adding a reader here and nothing else.

Supported today: .xlsx (openpyxl) and .csv.

Cells keep their original type: a numeric Excel cell stays a number, text stays
text. That distinction matters — ``12500.5`` from a numeric cell is already
unambiguous, while the *text* ``"12.500,50"`` needs locale rules to read.
Flattening everything to strings first would make the two indistinguishable.
"""
from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation

MAX_ROWS = 5000

Cell = str | int | float
Row = list[Cell]


class ReaderError(Exception):
    pass


def read_table(filename: str, content: bytes) -> tuple[list[str], list[Row]]:
    """Return (headers, rows). Raises ReaderError with a user-facing message."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        headers, rows = _read_xlsx(content)
    elif name.endswith((".csv", ".txt")):
        headers, rows = _read_csv(content)
    else:
        raise ReaderError(
            "Formato no soportado. Subí un archivo .xlsx o .csv. "
            "Si tenés un PDF, convertilo a Excel primero."
        )

    if not headers:
        raise ReaderError("El archivo no tiene una fila de encabezados.")
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise ReaderError("El archivo no tiene filas de datos.")
    if len(rows) > MAX_ROWS:
        raise ReaderError(
            f"El archivo tiene {len(rows)} filas y el máximo es {MAX_ROWS}. "
            "Dividilo en partes."
        )
    width = len(headers)
    rows = [(list(r) + [""] * width)[:width] for r in rows]
    return headers, rows


def _cell(value) -> Cell:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _read_xlsx(content: bytes) -> tuple[list[str], list[Row]]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface a readable message
        raise ReaderError(f"No se pudo leer el Excel: {exc}")
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    headers = [str(_cell(c)) for c in header_row]
    while headers and not headers[-1]:      # trailing unnamed columns are noise
        headers.pop()
    rows = [[_cell(c) for c in r][: len(headers)] for r in rows_iter]
    wb.close()
    return headers, rows


def _read_csv(content: bytes) -> tuple[list[str], list[Row]]:
    text = _decode(content)
    sample = text[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        # Spanish Excel exports use ';' — prefer it when clearly present.
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    all_rows = [
        [(c or "").strip() for c in r]
        for r in csv.reader(io.StringIO(text), delimiter=delimiter)
    ]
    if not all_rows:
        return [], []
    headers = all_rows[0]
    while headers and not headers[-1]:
        headers.pop()
    return headers, all_rows[1:]


def _decode(content: bytes) -> str:
    # utf-8-sig strips the BOM Excel writes; latin-1 is the usual fallback for
    # files from older Windows tooling and never raises.
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def parse_decimal(value: Cell, *, decimal_format: str = "es") -> Decimal:
    """Read a number from a cell.

    Numeric cells are already unambiguous and are used as-is. Text is parsed
    with the format the user picked at mapping time rather than guessed: with a
    single separator, "1.500" is genuinely ambiguous (1500 or 1.5) and guessing
    wrong silently corrupts a cost by a factor of 1000.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value))

    s = str(value or "").strip()
    if not s:
        raise ValueError("vacío")
    for junk in ("$", "ARS", "AR$", " ", " "):
        s = s.replace(junk, "")
    if decimal_format == "en":
        s = s.replace(",", "")
    else:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        raise ValueError(f'"{value}" no es un número válido')
